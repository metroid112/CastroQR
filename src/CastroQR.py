import pprint
import qrcode

from openpyxl import load_workbook

excel = load_workbook('test.xlsx')
sheet1 = excel.worksheets[0]
products = []
qr = []
for row in sheet1.iter_rows(min_row=2, max_col=4, values_only=True):
    products.append(row)
for product in products:
    property_string = ''
    for i, value in enumerate(product):
        if i == 0:
            property_string = f'ID: {value}'
        if i == 1:
            property_string += f' Stock: {value}'
        if i == 2:
            property_string += f' Descripción: {value}'
        if i == 3:
            property_string += f' Metros disponibles: {value}mt\u00b2'
    qr.append(property_string)
for i, q in enumerate(qr):
    pprint.pprint(q)
    img = qrcode.make(q).save(f'{i}.png', format='PNG')
pprint.pprint(qr)
